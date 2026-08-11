import pandas as pd
import openpyxl
import os
from datetime import datetime
import streamlit as st

def mostrar_modulo_compras(ruta_negocio):
    st.markdown("### 🛒 Módulo de Recepción de Compras y Control de Lotes")
    st.markdown("Registra la factura de tu proveedor, añade los productos y el sistema actualizará el inventario y las finanzas automáticamente.")

    archivo_base = os.path.join(ruta_negocio, "BASE DE DATOS.xlsx")
    archivo_compras = os.path.join(ruta_negocio, "Historial_Compras.xlsx")
    archivo_cuentas = os.path.join(ruta_negocio, "Cuentas_Por_Pagar.xlsx")
    archivo_gastos = os.path.join(ruta_negocio, "Registro_Gastos.xlsx")

    # Asegurar archivos base si no existen
    if not os.path.exists(archivo_compras):
        pd.DataFrame(columns=[
            'Fecha_Hora', 'Codigo', 'Descripcion', 'Proveedor', 'Numero_Factura',
            'Cantidad_Comprada', 'Costo_Unitario_Nuevo', 'Costo_Promedio_Final', 'Fecha_Vencimiento'
        ]).to_excel(archivo_compras, index=False)

    if not os.path.exists(archivo_cuentas):
        pd.DataFrame(columns=[
            'Proveedor', 'Numero_Factura', 'Fecha_Emision', 'Fecha_Vencimiento', 'Monto_Total', 'Estado'
        ]).to_excel(archivo_cuentas, index=False)

    if not os.path.exists(archivo_gastos):
        pd.DataFrame(columns=[
            'Fecha', 'Proveedor', 'Numero_Factura', 'Tipo_Pago', 'Categoria', 'Monto'
        ]).to_excel(archivo_gastos, index=False)

    if not os.path.exists(archivo_base):
        st.error(f"❌ Error crítico: No se encuentra el archivo maestro en '{archivo_base}'.")
        return

    # Cargamos la base maestra y proveedores
    wb = openpyxl.load_workbook(archivo_base)
    ws = wb.active
    df_base = pd.read_excel(archivo_base, dtype={'Código': str})
    headers = [cell.value for cell in ws[1] if cell.value is not None]

    col_stock = next((col for col in df_base.columns if str(col).strip().lower() in ['stock', 'cantidad', 'inventario', 'existencia']), 'Stock')
    col_costo = next((col for col in df_base.columns if 'costo' in str(col).lower() or 'compra' in str(col).lower()), 'Costo')
    col_venc = next((col for col in df_base.columns if 'vencimiento' in str(col).lower() or 'vence' in str(col).lower()), 'Fecha_Vencimiento')

    for col_name in [col_stock, col_costo, col_venc]:
        if col_name not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=col_name)
            headers.append(col_name)
            wb.save(archivo_base)

    idx_stock_ws = headers.index(col_stock) + 1
    idx_costo_ws = headers.index(col_costo) + 1
    idx_venc_ws = headers.index(col_venc) + 1

    # Intentamos cargar proveedores desde la base si existe la hoja
    try:
        df_prov_list = pd.read_excel(archivo_base, sheet_name="BD_Proveedores", dtype={'RUT': str})
        lista_proveedores = df_prov_list['Nombre'].tolist() if not df_prov_list.empty else ["Proveedor General"]
    except Exception:
        lista_proveedores = ["Proveedor General"]

    st.divider()
    st.markdown("#### 📄 1. Cabecera de la Factura")
   
    col_h1, col_h2, col_h3 = st.columns(3)
    with col_h1:
        proveedor_factura = st.selectbox("Nombre del Proveedor", options=lista_proveedores)
    with col_h2:
        num_factura = st.text_input("Número de Factura / Documento", value="FAC-001")
    with col_h3:
        condicion_pago = st.selectbox("Condición de Pago", ["Contado", "Crédito", "Cheque"])

    col_h4, col_h5 = st.columns(2)
    with col_h4:
        fecha_emision = st.date_input("Fecha de Emisión / Compra", value=datetime.today())
    with col_h5:
        fecha_vencimiento_factura = st.date_input("Fecha de Vencimiento (para Crédito / Cheque)", value=datetime.today())

    st.divider()
    st.markdown("#### 📦 2. Agregar Productos de la Compra")

    codigos_disponibles = df_base['Código'].astype(str).tolist() if 'Código' in df_base.columns else []
   
    # Usamos session_state para ir acumulando los items de esta compra antes de grabar
    if 'items_compra_actual' not in st.session_state:
        st.session_state.items_compra_actual = []

    with st.form("form_agregar_item_compra"):
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            codigo_prod = st.selectbox("Código / Producto", options=codigos_disponibles) if codigos_disponibles else st.text_input("Código")
        with c2:
            cant_comprada = st.number_input("Cantidad", min_value=0.0, value=1.0, step=1.0)
        with c3:
            costo_unit = st.number_input("Costo Unitario ($)", min_value=0.0, value=0.0, step=100.0)
        with c4:
            venc_lote = st.text_input("Vencimiento Lote (Ej: 2026-12-31)", value="Sin Vencimiento")

        btn_add = st.form_submit_button("➕ Añadir Producto a la Factura")
        if btn_add:
            if codigo_prod:
                match_p = df_base[df_base['Código'].astype(str).str.strip() == str(codigo_prod).strip()]
                desc_p = match_p['Descripción'].values[0] if not match_p.empty and 'Descripción' in match_p.columns else "Sin descripción"
               
                st.session_state.items_compra_actual.append({
                    'codigo': str(codigo_prod),
                    'descripcion': desc_p,
                    'cantidad': cant_comprada,
                    'costo': costo_unit,
                    'vencimiento': venc_lote,
                    'subtotal': cant_comprada * costo_unit
                })
                st.success(f"Agregado: {desc_p}")

    # Mostrar tabla temporal de lo que se va comprando
    if st.session_state.items_compra_actual:
        st.markdown("##### Productos en esta Factura:")
        df_temp = pd.DataFrame(st.session_state.items_compra_actual)
        st.dataframe(df_temp[['codigo', 'descripcion', 'cantidad', 'costo', 'subtotal', 'vencimiento']], use_container_width=True)
       
        monto_total_factura = df_temp['subtotal'].sum()
        st.markdown(f"### 💰 **Monto Total Factura: ${monto_total_factura:,.2f}**")

        col_acc1, col_acc2 = st.columns(2)
        with col_acc1:
            if st.button("🗑️ Limpiar / Cancelar Factura"):
                st.session_state.items_compra_actual = []
                st.rerun()
        with col_acc2:
            if st.button("💾 Guardar y Procesar Factura Completa", type="primary"):
                # PROCESAR GUARDADO DEFINITIVO
                compras_historial = []
               
                for item in st.session_state.items_compra_actual:
                    match = df_base[df_base['Código'].astype(str).str.strip() == item['codigo']]
                    if not match.empty:
                        row_index_df = match.index[0]
                       
                        val_stock = match[col_stock].values[0] if col_stock in match.columns else 0.0
                        stock_actual = float(val_stock) if pd.notna(val_stock) else 0.0

                        val_costo = match[col_costo].values[0] if col_costo in match.columns else 0.0
                        costo_anterior = float(val_costo) if pd.notna(val_costo) else 0.0

                        # Costo Promedio Ponderado
                        cant_n = item['cantidad']
                        costo_n = item['costo']
                        if (stock_actual + cant_n) > 0:
                            nuevo_cpp = ((stock_actual * costo_anterior) + (cant_n * costo_n)) / (stock_actual + cant_n)
                        else:
                            nuevo_cpp = costo_n

                        nuevo_stock = stock_actual + cant_n

                        # Actualizar DataFrame maestro
                        df_base.loc[row_index_df, col_stock] = nuevo_stock
                        df_base.loc[row_index_df, col_costo] = nuevo_cpp
                        df_base.loc[row_index_df, col_venc] = item['vencimiento']

                        # Actualizar Excel maestro
                        row_ws_idx = row_index_df + 2
                        ws.cell(row=row_ws_idx, column=idx_stock_ws, value=nuevo_stock)
                        ws.cell(row=row_ws_idx, column=idx_costo_ws, value=nuevo_cpp)
                        ws.cell(row=row_ws_idx, column=idx_venc_ws, value=item['vencimiento'])

                        compras_historial.append({
                            'Fecha_Hora': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'Codigo': item['codigo'],
                            'Descripcion': item['descripcion'],
                            'Proveedor': proveedor_factura,
                            'Numero_Factura': num_factura,
                            'Cantidad_Comprada': cant_n,
                            'Costo_Unitario_Nuevo': costo_n,
                            'Costo_Promedio_Final': nuevo_cpp,
                            'Fecha_Vencimiento': item['vencimiento']
                        })

                # Guardar base maestra
                for cell in ws['A']:
                    if cell.row > 1:
                        cell.number_format = '@'
                wb.save(archivo_base)

                # Guardar en Historial de Compras
                df_hist_ant = pd.read_excel(archivo_compras)
                df_hist_nue = pd.DataFrame(compras_historial)
                pd.concat([df_hist_ant, df_hist_nue], ignore_index=True).to_excel(archivo_compras, index=False)

                # 1. REGISTRAR SIEMPRE EN REGISTRO DE GASTOS
                df_gastos_ant = pd.read_excel(archivo_gastos)
                nuevo_gasto = pd.DataFrame([{
                    'Fecha': str(fecha_emision),
                    'Proveedor': proveedor_factura,
                    'Numero_Factura': num_factura,
                    'Tipo_Pago': condicion_pago,
                    'Categoria': 'Mercadería / Compras',
                    'Monto': monto_total_factura
                }])
                pd.concat([df_gastos_ant, nuevo_gasto], ignore_index=True).to_excel(archivo_gastos, index=False)

                # 2. SI ES CRÉDITO O CHEQUE, REGISTRAR EN CUENTAS POR PAGAR (Y CALENDARIO)
                if condicion_pago in ["Crédito", "Cheque"]:
                    df_cuentas_ant = pd.read_excel(archivo_cuentas)
                    nueva_cuenta = pd.DataFrame([{
                        'Proveedor': proveedor_factura,
                        'Numero_Factura': num_factura,
                        'Fecha_Emision': str(fecha_emision),
                        'Fecha_Vencimiento': str(fecha_vencimiento_factura),
                        'Monto_Total': monto_total_factura,
                        'Estado': 'Pendiente'
                    }])
                    pd.concat([df_cuentas_ant, nueva_cuenta], ignore_index=True).to_excel(archivo_cuentas, index=False)

                # Limpiar sesión y notificar éxito
                st.session_state.items_compra_actual = []
                st.success("🎉 ¡Factura procesada con éxito! Stock actualizado, registrada en gastos y en cuentas por pagar si correspondía.")
                st.rerun()
    else:
        st.info("ℹ️ Añade al menos un producto usando el formulario de arriba para armar la factura.")