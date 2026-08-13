import pandas as pd
import openpyxl
import os
from datetime import datetime
from supabase import create_client, Client

# --- CONFIGURACIÓN DE LA NUBE (SUPABASE) ---
# ⚠️ Copia tu URL y KEY exactamente como las tienes en app.py
SUPABASE_URL = "https://dmkjlcjrobszhwasrofc.supabase.co"
SUPABASE_KEY = "sb_publishable_uGVmMWz7T9aShxTMm_Vrgw_QFvRyTmH"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

print("🛒 Iniciando Terminal de Caja y Ventas con Escáner...")

# --- SEGURIDAD: SOLICITAMOS EL RUT AL INICIAR EL TURNO ---
print("\n🔒 Control de Seguridad")
RUT_NEGOCIO = input("🔑 Ingresa el RUT de este local (ej: 12345678-9) para iniciar turno: ").strip()

if not RUT_NEGOCIO:
    print("❌ Error: Debes ingresar un RUT válido para operar la caja.")
    exit()

archivo_base = os.path.join("clientes", RUT_NEGOCIO, "BASE DE DATOS.xlsx")

if not os.path.exists(archivo_base):
    print(f"❌ Error crítico: No se encuentra el archivo maestro '{archivo_base}'.")
else:
    # Cargamos el libro maestro local
    wb = openpyxl.load_workbook(archivo_base)
    ws = wb.active
    df_base = pd.read_excel(archivo_base, dtype={'Código': str})
    headers = [cell.value for cell in ws[1]]

    col_stock = next((col for col in df_base.columns if 'stock' in col.lower() or 'cantidad' in col.lower()), None)
    col_precio = next((col for col in df_base.columns if 'precio' in col.lower() or 'venta' in col.lower()), None)

    if not col_stock or not col_precio:
        print("⚠️ No se encontró la columna de Stock o Precio en la base maestra.")
    else:
        idx_stock_ws = df_base.columns.get_loc(col_stock) + 1
        
        print("\n---------------------------------------------------------")
        print(f"🟢 CAJA ABIERTA - LOCAL: {RUT_NEGOCIO}")
        print("---------------------------------------------------------")
        print("Instrucciones: Ingresa el código de barras del producto (o escribe 'salir' para cerrar caja).")

        carrito = []

        while True:
            codigo_ingresado = input("\nScan Código / Escribir Código: ").strip()

            if codigo_ingresado.lower() == 'salir':
                break

            if not codigo_ingresado:
                continue

            match = df_base[df_base['Código'].astype(str).str.strip() == codigo_ingresado]

            if match.empty:
                print(f"❌ Producto con código '{codigo_ingresado}' no encontrado en la base de datos.")
            else:
                row_index_df = match.index[0]
                descripcion = match['Descripción'].values[0] if 'Descripción' in match.columns else "Sin descripción"
               # --- EXTRACCIÓN SEGURA DE PRECIO ---
                try:
                    precio_venta = float(match[col_precio].values[0])
                except Exception:
                    precio_venta = 0.0
                
                # --- EXTRACCIÓN SEGURA DE STOCK ---
                try:
                    stock_actual = float(match[col_stock].values[0])
                except Exception:
                    # Si el Excel dice "Si" o tiene letras, asumimos que hay stock y le ponemos 99 para no frenar la venta
                    print(f"⚠️ Nota: Tu Excel dice '{match[col_stock].values[0]}' en vez de un número. Asumiendo stock disponible.")
                    stock_actual = 99.0
                print(f"📦 Producto: {descripcion}")
                print(f"💵 Precio Unitario: ${precio_venta:,.2f} | Stock Disponible: {stock_actual}")

                if stock_actual <= 0:
                    print("⚠️ ¡ALERTA! Producto sin stock disponible para la venta.")
                    continuar = input("¿Desea vender de todas formas? (s/n): ").strip().lower()
                    if continuar != 's':
                        continue

                try:
                    cantidad_comprada = float(input("Cantidad a vender (por defecto 1): ") or "1")
                except ValueError:
                    cantidad_comprada = 1

                nuevo_stock = stock_actual - cantidad_comprada
                df_base.loc[row_index_df, col_stock] = nuevo_stock

                row_ws_idx = row_index_df + 2 
                ws.cell(row=row_ws_idx, column=idx_stock_ws, value=nuevo_stock)

                total_linea = precio_venta * cantidad_comprada
                
                # --- PREPARAMOS EL DATO PARA LA NUBE ---
                carrito.append({
                    "rut_empresa": RUT_NEGOCIO,
                    "fecha": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    "detalle": f"{descripcion} (Cant: {cantidad_comprada})",
                    "monto": total_linea,
                    "metodo_pago": "Efectivo",
                    "documento": "Ticket de Venta"
                })

                print(f"✅ Agregado al ticket. Subtotal línea: ${total_linea:,.2f} | Nuevo Stock: {nuevo_stock}")

        # Al cerrar caja
        if carrito:
            for cell in ws['A']:
                if cell.row > 1:
                    cell.number_format = '@'
                    cell.data_type = 's'
            wb.save(archivo_base) 

            print("\n⏳ Subiendo ventas a la nube (Supabase)...")
            try:
                supabase.table("ventas").insert(carrito).execute()
                print("✅ ¡Ventas registradas con éxito en la nube!")
            except Exception as e:
                print(f"❌ Error al conectar con la nube: {e}")
                df_respaldo = pd.DataFrame(carrito)
                df_respaldo.to_excel("Respaldo_Ventas_Emergencia.xlsx", index=False)
                print("⚠️ Falló el internet. Las ventas se guardaron localmente en 'Respaldo_Ventas_Emergencia.xlsx'.")

            print("\n---------------------------------------------------------")
            print("🏁 CAJA CERRADA CON ÉXITO")
            print(f"📁 Stock local actualizado en '{archivo_base}'.")
            print("---------------------------------------------------------")
        else:
            print("\nℹ️ No se registraron ventas en esta sesión de caja.")