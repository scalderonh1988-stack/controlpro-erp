import pandas as pd
import openpyxl
import os

print("🔄 Iniciando módulo de actualización masiva de Precios y Stock...")

archivo_base = "BASE DE DATOS.xlsx"
archivo_actualizacion = "Actualizacion_Precios.xlsx"

# Verificamos que exista la base principal
if not os.path.exists(archivo_base):
    print(f"❌ Error crítico: No se encuentra el archivo maestro '{archivo_base}'.")
else:
    # Si aún no existe un archivo de actualización de prueba, lo creamos automáticamente para ti
    if not os.path.exists(archivo_actualizacion):
        print(f"⚠️ No se encontró '{archivo_actualizacion}'. Creando un archivo de ejemplo...")
        df_ejemplo = pd.DataFrame({
            'Código': ['7802810012531'], # Código de ejemplo (Aceite Canola)
            'Precio': [1890],               # Nuevo precio de venta
            'Stock': [24]                   # Nueva cantidad en inventario
        })
        df_ejemplo.to_excel(archivo_actualizacion, index=False)
        print(f"📁 Archivo de ejemplo creado: {archivo_actualizacion}. Puedes editarlo cuando quieras.")

    # 1. Cargamos el libro maestro con openpyxl para respetar formato y diseño original
    wb = openpyxl.load_workbook(archivo_base)
    ws = wb.active

    # 2. Leemos los datos con pandas para hacer el cruce de manera rápida y limpia
    df_base = pd.read_excel(archivo_base, dtype=str)
    df_act = pd.read_excel(archivo_actualizacion, dtype=str)

    # Identificamos las columnas de nuestra base maestra
    headers = [cell.value for cell in ws[1]]
    print(f"📊 Columnas detectadas en tu base maestra: {headers}")

    # Verificamos si existen columnas de Precio o Stock en la base maestra
    # (Si tus columnas se llaman distinto, adaptaremos esto en un segundo)
    actualizados = 0

    # Recorremos el archivo de actualización
    for _, row_act in df_act.iterrows():
        cod_act = str(row_act['Código']).strip()
        
        # Buscamos la fila correspondiente en la hoja de Excel maestra
        for row_idx in range(2, ws.max_row + 1):
            cod_base = str(ws.cell(row=row_idx, column=1).value).strip()
            
            if cod_base == cod_act:
                # Si el código coincide, actualizamos las columnas si vienen en el archivo
                for col_name in df_act.columns:
                    if col_name in headers:
                        col_idx = headers.index(col_name) + 1
                        nuevo_valor = row_act[col_name]
                        ws.cell(row=row_idx, column=col_idx, value=nuevo_valor)
                actualizados += 1
                break

    # Blindamos estrictamente la columna A (Códigos) para que el escáner nunca falle
    for cell in ws['A']:
        if cell.row > 1:
            cell.number_format = '@'
            cell.data_type = 's'

    # Guardamos los cambios en la base de datos maestra
    wb.save(archivo_base)
    print(f"✅ ¡Actualización masiva completada con éxito! Se actualizaron {actualizados} productos.")
    print(f"📁 Tu archivo '{archivo_base}' está al día y con los códigos blindados.")